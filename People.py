# -*- coding: utf-8 -*-
class People(object):

    def __init__(self):
        self.people = dict()

    def loadFromXLS(self,workbook, worksheet):
        import openpyxl
        wb = openpyxl.load_workbook(workbook)
        sh = wb.get_sheet_by_name(worksheet)
        count_columns = 0
        for i in range(1, 100):
            if sh.cell(row=1, column=i).value:
                # print (sh_history.cell(row=1,column=i).value)
                count_columns += 1
            else:
                break
        p_heads = list()
        for line in sh:
            if line[0].value:
                if len(p_heads) == 0:
                    tmp_heads = list(x.value for x in line if x.col_idx <= count_columns)
                    p_heads = tmp_heads[:5]
                    continue
                else:
                    tmp_values = list(x.value for x in line if x.col_idx <= count_columns)
                    p_values = tmp_values[:5]
                    p_name = p_values[2]

                self.people[p_name] = dict()
                for h, v in list(zip(p_heads, p_values)):
                    if 'Фамилия и Имя сотрудника' not in h:
                        if v:
                            self.people[p_name][h] = v
                self.people[p_name]["Trainings"] = list()
                for h, v in list(zip(tmp_heads[5:], tmp_values[5:])):
                    if v: self.people[p_name]["Trainings"].append(h)

    def removeIfTRainigIsInOtherList (self,otherList):
        deleted=list()
        for prt in self.people:
            for tr in self.people[prt]["Trainings"]:
                if prt in otherList:
                    if tr in otherList[prt]["Trainings"]:
                        deleted.append(prt + ': ' +tr)
                        self.people[prt]["Trainings"].remove(tr)
        return deleted

    def deleteIfDependenceIsNotInHistry(self,history,rules):
        deleted=list()
        for prt in self.people:
            self.people[prt]["Sort"] = list()
            for tr in self.people[prt]["Trainings"]:
                prevs = rules.getPreviousTraining(tr)
                if not prevs: continue
                for prev in prevs:
                    if prt in history.people:
                        if prev not in history.people[prt]["Trainings"]:
                            if prev not in self.people[prt]["Trainings"]:
                                tmp_l = self.people[prt]["Trainings"]

                                tmp_l.remove(tr)
                                self.people[prt]["Trainings"] = tmp_l
        return deleted

    def makeListOfTrainings(self):
        self.listOfTrainings = list()
        for prt in self.people:
            for tr in self.people[prt]["Trainings"]:
                self.listOfTrainings.append(str(prt)+'|'+tr)

    def countOfListOfTrainings(self):
        return len(self.listOfTrainings)

    def getRole(self, participant):
        return self.people[participant]['Должность']

    def getRegion(self,participant):
        return self.people[participant]['регион']

    def hasParticipantTraining(self,participant,training):
        res = False
        for tr in self.people[participant]['Trainings']:
            if tr==training:
                res = True
        return res

    def hasParticipantTraining2(self,participant,training):
        res = False
        if participant+'|'+training in self.listOfTrainings:
            res = True
        return res

    def isPreviousInHistory(self,participant,previous):
        res = False
        count_is = 0
        if len(previous)>0:
            for pr in previous:
                if pr in self.people[participant]['Trainings']:
                    count_is +=1
        else:
            res = True
        if count_is == len(previous):
            res = True
        return res
