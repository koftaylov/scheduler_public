# -*- coding: utf-8 -*-
"""
reading History Excel file
"""


def load_participants(workbook, worksheet):
    import openpyxl
    wb = openpyxl.load_workbook(workbook)
    sh = wb.get_sheet_by_name(worksheet)
    count_columns = 0
    for i in range(1, 100):
        if sh.cell(row=1, column=i).value:
            # print (sh_history.cell(row=1,column=i).value)
            count_columns += 1
        else:
            break
    p_heads = list()
    participants = dict()
    for line in sh:
        if line[0].value:
            if len(p_heads) == 0:
                tmp_heads =list(x.value for x in line if x.col_idx <= count_columns)
                p_heads = tmp_heads[:5]
                # print("heads: ", p_heads)
                continue
            else:
                tmp_values = list(x.value for x in line if x.col_idx <= count_columns)
                p_values = tmp_values[:5]
                # print("values: ", p_values)
                p_name = p_values[2]

            participants[p_name] = dict()
            for h, v in list(zip(p_heads, p_values)):
                # print ('p_name: ',p_name,'. h: ',h,'. v: ',v)
                # print(p_name.__class__)
                if 'Фамилия и Имя сотрудника' not in h:
                    if v:
                        participants[p_name][h] = v
            participants[p_name]["Trainings"] = list()
            for h,v in list(zip(tmp_heads[5:], tmp_values[5:])):
                # if v: participants[p_name]["Тренинги"][h]=v
                if v: participants[p_name]["Trainings"].append(h)

    return participants
    # print(history)


def load_participants_last(workbook, worksheet):
    import openpyxl
    wb = openpyxl.load_workbook(workbook)
    sh = wb.get_sheet_by_name(worksheet)
    count_columns = 0
    for i in range(1, 100):
        if sh.cell(row=1, column=i).value:
            # print (sh_history.cell(row=1,column=i).value)
            count_columns += 1
        else:
            break
    p_heads = list()
    participants = dict()
    for line in sh:
        if line[0].value:
            if len(p_heads) == 0:
                p_heads = list(x.value for x in line if x.col_idx <= count_columns)
                # print("heads: ", p_heads)
                continue
            else:
                p_values = list(x.value for x in line if x.col_idx <= count_columns)
                # print("values: ", p_values)
                p_name = p_values[2]

            participants[p_name] = dict()
            for h, v in list(zip(p_heads, p_values)):
                # print ('p_name: ',p_name,'. h: ',h,'. v: ',v)
                # print(p_name.__class__)
                if 'Фамилия и Имя сотрудника' not in h:
                    if v:
                        participants[p_name][h] = v
    return participants
    # print(history)

# -------------------------------------------------------------------------------------------------------------------

def load_rules(filename):
    rules = dict()
    input = open(filename, 'r')
    while True:
        line = input.readline()

        # Trainings
        if line.split(';')[0] == "Trainings":
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[1] or line.split(";")[2]:
                if '\n' in line: line = line.replace('\n', '')
                if line.split(";")[1]:
                    sub_head = line.split(';')[1]
                    rules[head][sub_head] = dict()
                    keys = line.split(';')[3:]
                else:
                    val = line.split(';')[2]
                    values = line.split(';')[3:]

                    rules[head][sub_head][val] = dict(zip(keys, values))
                    pr = rules[head][sub_head][val]["previous"]
                    if pr != "":
                        rules[head][sub_head][val]["previous"] = list(pr.split("|"))
                    else:
                        rules[head][sub_head][val]["previous"] = list()
                line = input.readline()
        # Trainings end

        line = input.readline()
        # roles
        if line.split(';')[0] == 'roles':
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[0]:

                if '\n' in line: line = line.replace('\n', '')
                val = line.split(';')[0]
                values = line.split(';')[1:]
                rules[head][val] = [value for value in values if value != '']
                line = input.readline()

        # roles end

        line = input.readline()
        # Region_Manager_roles
        if line.split(';')[0] == 'Region_Manager_roles':
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[0]:

                if '\n' in line: line = line.replace('\n', '')
                val = line.split(';')[0]
                values = line.split(';')[1:]
                rules[head][val] = [value for value in values if value != '']
                line = input.readline()

        # Region_Manager_roles end

        line = input.readline()

        # roles_hierarchy
        if line.split(';')[0] == 'roles_hierarchy':
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[0]:

                if '\n' in line: line = line.replace('\n', '')
                val = line.split(';')[0]
                values = int(line.split(';')[1])
                rules[head][val] = values
                line = input.readline()
        # roles_hierarchy end

        line = input.readline()
        # gap start
        if line.split(';')[0] == 'gap':
            head = line.split(';')[0]

            gap = input.readline().split(";")[0]
            rules[head] = int(gap)
        # gap end
        line = input.readline()
        line = input.readline()

        # region_max_amount start
        if line.split(';')[0] == 'region_max_amount':
            head = line.split(';')[0]
            region_max_amount = int(input.readline().split(";")[0])
            rules[head] = region_max_amount

        # region_max_amount end
        line = input.readline()
        line = input.readline()
        # workdays start
        if line.split(';')[0] == 'workdays':
            head = line.split(';')[0]
            line = input.readline()
            if '\n' in line: line = line.replace('\n', '')
            rules[head] = [int(value) for value in line.split(';') if value != '']

        # workdays end

        line = input.readline()
        line = input.readline()
        # holidays start
        if line.split(';')[0] == 'holidays':
            from datetime import date
            from datetime import timedelta
            head = line.split(';')[0]
            # holidays = input.readline().split(";")[0]
            rules[head] = list()
            line = input.readline()
            holidays = list()
            while line.split(";")[0]:
                if '\n' in line: line = line.replace('\n', '')
                start_day = date(int(line.split(';')[0].split('-')[2]), int(line.split(';')[0].split('-')[1]),
                                 int(line.split(';')[0].split('-')[0]))

                # holidays.append(start_day.isoformat())
                holidays.append(start_day)
                if line.split(';')[1]:
                    end_day = date(int(line.split(';')[1].split('-')[2]), int(line.split(';')[1].split('-')[1]),
                                   int(line.split(';')[1].split('-')[0]))
                    for x in range(int((end_day - start_day).days)):
                        # holidays.append((start_day + timedelta(days=x + 1)).isoformat())
                        holidays.append((start_day + timedelta(days=x + 1)))
                rules[head] = holidays

                line = input.readline()

        # holidays end
        line = input.readline()
        # print(line)
        # holidays_trainers start
        if line.split(';')[0] == "holidays_trainers":
            from datetime import date
            from datetime import timedelta
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[1] or line.split(";")[2]:
                if '\n' in line: line = line.replace('\n', '')
                if line.split(";")[1]:
                    sub_head = line.split(';')[1]
                    rules[head][sub_head] = list()
                    # keys = line.split(';')[3:]
                else:
                    start_day = date(int(line.split(';')[2].split('-')[2]),
                                     int(line.split(';')[2].split('-')[1]),
                                     int(line.split(';')[2].split('-')[0]))
                    holidays_trainers = list()
                    # holidays_trainers.append(start_day.isoformat())
                    holidays_trainers.append(start_day)
                    if line.split(';')[3]:
                        end_day = date(int(line.split(';')[3].split('-')[2]),
                                       int(line.split(';')[3].split('-')[1]),
                                       int(line.split(';')[3].split('-')[0]))
                        for x in range(int((end_day - start_day).days)):
                            # holidays_trainers.append((start_day + timedelta(days=x + 1)).isoformat())
                            holidays_trainers.append((start_day + timedelta(days=x + 1)))
                    rules[head][sub_head] = holidays_trainers
                line = input.readline()
        # holidays_trainers end
        line = input.readline()
        # trainer_max_week
        if line.split(';')[0] == 'trainer_max_week':
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[0]:

                if '\n' in line: line = line.replace('\n', '')
                val = line.split(';')[0]
                values = int(line.split(';')[1])
                rules[head][val] = values
                line = input.readline()
        # trainer_max_week end
        line = input.readline()

        # trainer_week_gap
        if line.split(';')[0] == 'trainer_week_gap':
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[0]:

                if '\n' in line: line = line.replace('\n', '')
                val = line.split(';')[0]
                values = int(line.split(';')[1])
                rules[head][val] = values
                line = input.readline()

        # trainer_week_gap end

        line = input.readline()

        if line.split(';')[0] == 'inadmissible_trainings':
            head = line.split(';')[0]
            rules[head] = dict()
            line = input.readline()
            while line.split(";")[0]:

                if '\n' in line: line = line.replace('\n', '')
                val = line.split(';')[0]
                values = line.split(';')[1:]
                rules[head][val] = [value for value in values if value != '']
                line = input.readline()

        break

    # print (rules)
    # print(rules.keys())
    # for x in rules: print(x, rules[x])
    input.close()

    """for trainer in rules['Trainings']:
        list_of_training =list(rules['Trainings'][trainer].keys())
        for training in list_of_training:
            if rules['Trainings'][trainer][training]['start_day'] == '':
                del rules['Trainings'][trainer][training]
                print('del',training, 'from', trainer)
    """
    return rules
# ----------------------------------------------------------------------------------------


def printDict(dict_for_print, prfx = "\t", iter = 0):
    if dict_for_print.__class__ != dict().__class__: return
    if iter.__class__ != int().__class__: return
    if prfx.__class__ != str().__class__: return

    for k in dict_for_print:
        if dict_for_print[k].__class__ == dict().__class__:
            tmp_dict = dict_for_print[k]
            print(prfx*iter, k, sep="")
            printDict(tmp_dict, prfx, iter=iter+1)
        else:
            print(prfx*iter, k, ": ", dict_for_print[k], sep="")
# ------------------------------------------------------------------------------------------


def getPreviousTraining (training, rules):
    if not training: return
    for trainer in rules["Trainings"]:
        if training in rules["Trainings"][trainer]:
            if rules["Trainings"][trainer][training]["previous"]:
                return rules["Trainings"][trainer][training]["previous"]
            else: continue
    return
# -------------------------------------------------------------------------------------------


def printPrtsForTrainingsCount(prts_for_trainings):
    total = 0
    for tr in prts_for_trainings:
        total += len(prts_for_trainings[tr])
        print(tr, ": ", len(prts_for_trainings[tr]))
    print("Total: ", total)
# --------------------------------------------------------------------------------------------
# проверка группы на соответствие правилам
# для каждого сотрудника определить причины невозможности быть в группе
def checkGroup(group):
    pass


def CheckParticipationPossibility(week, participant):

    poss = true
    comment = ""

    # check1: roles
    participant_role = demand[participant]["Должность"]
    possible_roles = rules["roles"][participant_role]

    for for_training in cal_for_tr_1.scheduleweeks[week]:
        print(for_training)
        for other_participants in cal_for_tr_1.scheduleweeks[week][for_training]:
            print(other_participants)
            if demand[other_participants]["Должность"] not in possible_roles:
                poss = false
                comment = "roles"
                # return poss, comment
    # check2: pouse
    for occupiedweek in cal_for_tr_1.participantweeks[participant]:
        pouse = abs(occupiedweek - week)
        if pouse < 7:
            poss = false
            comment = "pouse"
    return poss, comment


def AddParticipantToTraining(week, participant):

    # add to cal_for_tr_1.participantweeks[participant]
    cal_for_tr_1.participantweeks[participant][week] = cal_for_tr_1.scheduleweeks[week]


