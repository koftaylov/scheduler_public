# -*- coding: utf-8 -*-
class TrainerCalendar(object):

    def __init__(self, cal_year, trainer,rules):
        # trainer#1, trainer#2
        import datetime
        dd = datetime.date(cal_year, 1, 1)
        nxt = datetime.timedelta(days=1)
        cln = list()
        while dd.year == 2019: #comment koftaylov 201912170006
            cln.append(dd)
            dd += nxt
            cln[:] = [x for x in cln if x not in rules["holidays"] and x.isoweekday() in rules["workdays"]]
            self.cln_trainer = cln.copy()
            self.cln_trainer[:] = [x for x in self.cln_trainer if x not in rules["holidays_trainers"][trainer]]
        #end
        self.freeweeks = set(map(lambda x: x.isocalendar()[1], self.cln_trainer))
        self.scheduleweeks = dict.fromkeys(self.freeweeks)
        self.participantweeks = dict()
        i=1
        self.participantslist = list()
        self.trainings=dict()

        list_days=list()
        list_days.append(self.cln_trainer[0])
        prev=self.cln_trainer[0]
        for day in self.cln_trainer[1:]:
            if day-prev>datetime.timedelta(days=1):
                self.trainings[i]=dict()
                self.trainings[i]['start_day']=list_days[0]
                self.trainings[i]['count_days']=len(list_days)
                self.trainings[i]['training']=''
                self.trainings[i]['list']=list()
                self.trainings[i]['listOfRoles'] = list()
                list_days.clear()
                i+=1
            list_days.append(day)
            prev=day
            if day==self.cln_trainer[-1]:
                self.trainings[i] = dict()
                self.trainings[i]['start_day'] = list_days[0]
                self.trainings[i]['count_days'] = len(list_days)
                self.trainings[i]['training'] = ''
                self.trainings[i]['list'] = list()
                self.trainings[i]['listOfRoles'] = list()
                list_days.clear()
                i += 1

    def printweeks(self):
        print(self.freeweeks)

    def printself(self):
        print(self.cln_trainer)

    def printiso(self):
        print(list(map(lambda x: x.isoformat(), self.cln_trainer)))

    def printisocalendar(self):
        print(list(map(lambda x: x.isocalendar(), self.cln_trainer)))

    def printweeknums(self):
        print(list(map(lambda x: x.isocalendar()[1], self.cln_trainer)))

    def get(self):
        return self.cln_trainer

    def amountfreeweeks(self):
        tr_week = 1
        pr = 0
        for x in self.cln_trainer:
            if pr > x.isoweekday():
                tr_week += 1
            pr = x.isoweekday()
        return tr_week

    def copy(self):
        return self.cln_trainer

    def isTrainingInCalendar(self,training):
        res=False
        for num in self.trainings:
            if self.trainings[num]['training']==training:
                res = True
                break
        return res

    def addParticipantToTraining(self,participant,training,demand,rules,history):
        from datetime import timedelta
        res = False
        role = demand.getRole(participant)
        region = demand.getRegion(participant)
        for num in self.trainings:
            if self.trainings[num]['training'] == training:

                if ((True if int(rules.rules['Trainings']['trainer#1'][training]['roles_rules'])==0
                    else rules.isRegionManagerPossible(participant,self.trainings[num]['list'],demand,training)
                    if int(rules.rules['Trainings']['trainer#1'][training]['roles_rules'])==1 else rules.isRolePossible(role,self.trainings[num]['listOfRoles']))
                and rules.isMaxParticipantsPossible(len(self.trainings[num]['list']),training)
                and self.isGapPossible(self.trainings[num]['start_day'],participant,int(rules.rules['gap']))
                and rules.isMaxFromRegionPossible(participant,self.trainings[num]['list'],demand)
                and (self.isPreviousInCalendar(num,participant,rules.rules['Trainings']['trainer#1'][training]['previous'])
                        or self.isPreviousInHistory(participant,rules.rules['Trainings']['trainer#1'][training]['previous'],history))
                and res == False):
                    self.trainings[num]['list'].append(participant)
                    self.trainings[num]['listOfRoles'].append(demand.people[participant]['Должность'])
                    self.participantslist.append(participant+'|'+training)
                    self.trainings[num]['end_day']=self.trainings[num]['start_day']+timedelta(
                        days=int(rules.rules['Trainings']['trainer#1'][training]['days']))
                    res = True

        if not (res):
            for num in self.trainings:
                if (self.trainings[num]['training'] =='' and not (res)
                and self.isGapPossible(self.trainings[num]['start_day'],participant,int(rules.rules['gap']))
                and (self.isPreviousInCalendar(num,participant,rules.rules['Trainings']['trainer#1'][training]['previous'])
                     or self.isPreviousInHistory(participant,rules.rules['Trainings']['trainer#1'][training]['previous'],history))
                and rules.isPossibleForCalendarDates(training,self.trainings[num]['start_day'])  # Start day and End day for Training
                ):
                    self.trainings[num]['training'] = training
                    self.trainings[num]['list'].append(participant)
                    self.trainings[num]['listOfRoles'].append(demand.people[participant]['Должность'])
                    self.trainings[num]['end_day'] = self.trainings[num]['start_day'] + timedelta(
                        days=int(rules.rules['Trainings']['trainer#1'][training]['days']))
                    self.participantslist.append(participant + '|' + training)
                    res = True
                    continue
        return res

    def isPreviousInHistory(self,participant,previous,history):
        res = False
        if len(previous)==0:
            res = True
        else:
            count_is = 0
            for pr in previous:
                if participant in history.people and pr in history.people[participant]['Trainings']:
                    count_is+=1
            if count_is == len(previous):
                res=True
        return res

    def isPreviousInCalendar(self,num,participant,previous):
        res = False
        if len(previous)==0:
            res= True
        else:
            for pr in previous:
                for n in self.trainings:
                    if self.trainings[n]['training']==pr and n<=num:
                        if participant in self.trainings[n]['list']:
                            res = True
        return res

    def isGapPossible(self,date,participant,gapInDays):
        res = True
        import datetime
        listOfWeeks = list()
        for num in self.trainings:
            if participant in self.trainings[num]['list']:
                listOfWeeks.append(self.trainings[num]['start_day'])
        if len(listOfWeeks)>0:
            for d in listOfWeeks:
                if d>date:
                    if (d - date) < datetime.timedelta(days=gapInDays):
                        res = False
                else:
                    if (date-d) < datetime.timedelta(days=gapInDays):
                        res = False

        return res

    def saveToXLS(self,demand,best_score,total_trainings):
        print('Saving to XLSX')
        import datetime
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        from openpyxl.styles import PatternFill
        from openpyxl.styles import Font
        wb=load_workbook("schedule.xlsm")
        sh=wb.get_sheet_by_name("TrainingPlan")
        sh.title='Training Plan'
        listTrainings=list()
        for num in self.trainings:
            if self.trainings[num]["training"]!='':
                if self.trainings[num]["training"] not in listTrainings:
                    listTrainings.append(self.trainings[num]["training"])
                y=7+listTrainings.index(self.trainings[num]["training"])
                x=2+int(self.trainings[num]['start_day'].isocalendar()[1])
                sh['A'+str(y)]=self.trainings[num]["training"]
                sh[str(get_column_letter(x))+str(y)]=1
                comment = Comment('Trainer#1 \n ' + str(self.trainings[num]['start_day'].strftime('%d %b %Y'))+' - '+str(self.trainings[num]['end_day'].strftime('%d %b %Y')),
                                  'mr.scheduler')
                sh[str(get_column_letter(x)) + str(y)].comment=comment
                sh['B'+str(y)]=str(self.trainings[num]['end_day']-self.trainings[num]['start_day'])[0]
                sh[str(get_column_letter(x)) + str(y)].fill=PatternFill(bgColor="FFE2B552",fill_type="solid")

        waitinglist = dict()
        for prt in demand.listOfTrainings:
            if prt not in self.participantslist:
                tr = prt.split('|')[1]
                if tr in waitinglist:
                    waitinglist[tr].append(prt.split("|")[0])
                else:
                    waitinglist[tr] = list()
                    waitinglist[tr].append(prt.split("|")[0])
        # print(waitinglist)

        for training in listTrainings:

            last=0
            sh=wb.create_sheet(training)

            #link_to=wb['Training Plan']['A1'].value
            #link_from = wb[training]['G1']
            #link_from.value='plan'
            #link_from.hyperlink=hyperlink()
            #link_from.font=Font(color='FF4393EE')


            for num in self.trainings:
                if self.trainings[num]['training']==training:
                    sh["A"+str(last+1)]="Тренинг"
                    sh["B" + str(last+1)] = training
                    sh["A"+str(last+2)]="Дата тренинга"
                    sh["B" + str(last + 2)] = str(self.trainings[num]['start_day'].strftime('%d %b %Y'))
                    sh["A" + str(last + 3)] ="Место"

                    sh["A" + str(last + 5)] = "№"
                    sh["B" + str(last + 5)] = "Город"
                    sh["C" + str(last + 5)] = "ФИО"
                    sh["D" + str(last + 5)] = "Должность"
                    sh["E" + str(last + 5)] = "Линейный руководитель"
                    sh["F" + str(last + 5)] = "Регион"
                    y = 0
                    for participant in self.trainings[num]['list']:
                        y=y+1
                        sh["A" + str(last + 5 + y)] = y
                        sh["B" + str(last + 5 + y)] = demand.people[participant]['Город']
                        sh["C" + str(last + 5 + y)] = participant
                        sh["D" + str(last + 5 + y)] = demand.people[participant]['Должность']
                        sh["E" + str(last + 5 + y)] = demand.people[participant]['Линейный руководитель']
                        sh["F" + str(last + 5 + y)] = demand.people[participant]['регион']
                    last=last + 5 + y + 2

                    if training in waitinglist and len(waitinglist[training]) > 0:

                        sh["I" + str(1)] = "№"
                        sh["J" + str(1)] = "Город"
                        sh["K" + str(1)] = "ФИО"
                        sh["L" + str(1)] = "Должность"
                        sh["M" + str(1)] = "Линейный руководитель"
                        sh["N" + str(1)] = "Регион"
                        y = 0
                        for participant in waitinglist[training]:
                            y = y + 1
                            sh["I" + str(1 + y)] = y
                            sh["J" + str(1 + y)] = demand.people[participant]['Город']
                            sh["K" + str(1 + y)] = participant
                            sh["L" + str(1 + y)] = demand.people[participant]['Должность']
                            sh["M" + str(1 + y)] = demand.people[participant]['Линейный руководитель']
                            sh["N" + str(1 + y)] = demand.people[participant]['регион']





        pref=datetime.datetime.now()
        fname='Schedule_'+pref.strftime("%Y-%m-%d-%H-%M-%S")+' '+str(best_score)+' of '+str(total_trainings)  +'.xlsx'
        #wb.save("schedule_"+str(pref.year)+str(pref.month)+str(pref.day)+str(pref.hour)+str(pref.minute)+str(pref.second)+' ' + str(best_score) + '.xlsx')
        wb.save(fname)
        print('File saved: '+fname)
        return fname


    def deleteSmallTrainings(self,rules):
        for num in self.trainings:
            # print(str(num)+' | '+str(copy_cal_for_tr_1.trainings[num]['training'])+' | '+str(rules.rules['Trainings']['trainer#1'][copy_cal_for_tr_1.trainings[num]['training']]['min']))

            if self.trainings[num]['training'] != '' and (int(
                    rules.rules['Trainings']['trainer#1'][self.trainings[num]['training']]['min']) > len(
                    self.trainings[num]['list'])
                    ):
                for prt in self.trainings[num]['list']:
                    self.participantslist.remove(prt + '|' + self.trainings[num]['training'])
                self.trainings[num]['list'] = []
                self.trainings[num]['listOfRoles'] = []
                self.trainings[num]['training'] = ''
            else:
                pass

    def deleteUnfoldPeopleFromTraining(self,rules,demand):
        res = False
        res2 = dict()
        for num in self.trainings:
            if self.trainings[num]['training'] != '':
                amount_for_delete = len(self.trainings[num]['list']) % int(
                rules.rules['Trainings']['trainer#1'][self.trainings[num]['training']]['fold'])
                if amount_for_delete != 0:
                    for i in range(1,amount_for_delete+1):
                        res = False
                        list_hierarhy = [x for x in rules.rules['roles_hierarchy'].keys()]
                        list_hierarhy.reverse()
                        for  rule_participant in list_hierarhy:

                            for participant in  self.trainings[num]['list']:

                                if str(demand.people[participant]['Должность']).lower()==rule_participant:
                                    self.trainings[num]['list'].remove(participant)
                                    # print('delete unfold: ' + str(participant) + ': ' + str(demand.people[participant]['Должность'])) if printflag else None
                                    res = True
                                    break
                            if res == True:
                                break
                amount_for_delete = len(self.trainings[num]['list']) % int(
                    rules.rules['Trainings']['trainer#1'][self.trainings[num]['training']]['fold'])
                if amount_for_delete != 0:
                    res2[num] = False

        return res2
